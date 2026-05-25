#!/usr/bin/env python3
"""
Relatório de Assertividade por Copywriter e Categoria — IMPERA
Cruza ClickUp (produção + tipo + copywriter) com RedTrack (performance + validação).

Lógica: cada versão individual que rodou em tráfego = 1 criativo testado.
Atribuição ao copywriter que ESCREVEU aquela versão (match por range no ClickUp).

Uso:
  python3 relatorio_assertividade.py                         # Mês anterior inteiro
  python3 relatorio_assertividade.py 01/04 30/04            # Período custom DD/MM
  python3 relatorio_assertividade.py 01/04/2026 30/04/2026  # Com ano

Saída: PDF + Excel em ~/Documents/

Crontab sugerido: dia 3 de cada mês (mês anterior completo)
  3 9 3 * * cd ~/Scripts && python3 relatorio_assertividade.py 2>> logs/assertividade.log
"""

import sys, os, re, json, time as _time
from collections import defaultdict
from datetime import datetime, timedelta
import calendar

sys.path.insert(0, os.path.expanduser("~/Scripts"))
from relatorio_performance_criativos import (
    extract_creative_id, parse_cu_task, find_copywriter,
    clean_ag_name, parse_camp, COPY_LIST, TRAFEGO_LIST, REDTRACK_KEY,
    RIP_COPY
)
from impera_cache import cached_cu_tasks, rt_fetch_single

# === CONFIG ===
ROAS_MIN = 1.8
CPA_META = 180
SPEND_MINIMO = 50
RATE_LIMIT_DELAY = 1.5
MAX_RETRIES = 5

CATEGORIAS = [
    "Criativo Novo", "Variação Vídeo", "Imagem Nova",
    "Variação Imagem", "Ripagem"
]

OUTPUT_DIR = os.path.expanduser("~/Documents")


# === PARSE DATES ===
def parse_dates():
    if len(sys.argv) >= 3:
        d1, d2 = sys.argv[1], sys.argv[2]
        for fmt in ("%d/%m/%Y", "%d/%m"):
            try:
                dt1 = datetime.strptime(d1, fmt)
                dt2 = datetime.strptime(d2, fmt)
                if fmt == "%d/%m":
                    dt1 = dt1.replace(year=datetime.now().year)
                    dt2 = dt2.replace(year=datetime.now().year)
                return dt1.strftime("%Y-%m-%d"), dt2.strftime("%Y-%m-%d")
            except ValueError:
                continue
        print(f"Erro: formato de data inválido. Use DD/MM ou DD/MM/YYYY")
        sys.exit(1)
    else:
        today = datetime.now()
        first_this_month = today.replace(day=1)
        last_month_end = first_this_month - timedelta(days=1)
        last_month_start = last_month_end.replace(day=1)
        return last_month_start.strftime("%Y-%m-%d"), last_month_end.strftime("%Y-%m-%d")


# === REDTRACK FETCH (robusto) ===
def fetch_rt_all_adgroups(date_from, date_to):
    print("  Buscando campanhas ativas...", flush=True)
    campaigns = rt_fetch_single({
        "api_key": REDTRACK_KEY, "group": "campaign",
        "date_from": date_from, "date_to": date_to, "per": "500",
    })
    active = [c for c in campaigns if float(c.get("cost", 0)) > 0]
    print(f"  {len(active)} campanhas ativas", flush=True)

    all_rows = []
    failed = []

    for i, camp in enumerate(active):
        cid = camp["campaign_id"]
        cname = camp.get("campaign", "")
        cost = float(camp.get("cost", 0))

        success = False
        for attempt in range(MAX_RETRIES):
            try:
                rows = rt_fetch_single({
                    "api_key": REDTRACK_KEY, "group": "rt_adgroup",
                    "campaign_id": cid, "date_from": date_from,
                    "date_to": date_to, "per": "1000",
                })
                for r in rows:
                    r["campaign_id"] = cid
                    r["campaign"] = cname
                all_rows.extend(rows)
                success = True
                break
            except Exception as e:
                if "429" in str(e):
                    _time.sleep(RATE_LIMIT_DELAY * (2 ** (attempt + 1)))
                elif attempt < MAX_RETRIES - 1:
                    _time.sleep(RATE_LIMIT_DELAY * 2)
                else:
                    failed.append({"campaign": cname, "cost": cost, "error": str(e)})

        if not success:
            print(f"  ⚠️  FALHA: {cname[:50]} (R${cost:.0f})", flush=True)
        _time.sleep(RATE_LIMIT_DELAY)
        if (i + 1) % 10 == 0:
            print(f"  ... {i+1}/{len(active)} campanhas", flush=True)

    print(f"  {len(all_rows)} adgroups | {len(failed)} falhas", flush=True)
    return all_rows, failed


# === CLASSIFICAÇÃO ===
def classify_by_context(base_id, version, nicho, tasks_list):
    """Classifica pela tarefa no ClickUp que contém essa versão."""
    if re.match(r"^(CE|CY|CC)\d+$", base_id):
        if not version:
            return "Ripagem"
        return "Variação Vídeo"
    if re.match(r"^C\d+$", base_id) and not version:
        return "Ripagem"

    for t in tasks_list:
        if nicho and t["nicho"] and t["nicho"] != nicho:
            continue
        if base_id not in t.get("all_ids", []):
            continue

        task_name = t["name"].upper()
        is_img = "-IMG" in task_name or bool(re.search(r"\bIMG\d+", task_name))
        has_ad_version = bool(re.search(r"AD\d+V\d+", task_name))
        v_match = re.search(r"\[V(\d+)(?:\s*-\s*V?(\d+))?\]", task_name)

        if is_img:
            if v_match:
                v_start = int(v_match.group(1))
                v_end = int(v_match.group(2)) if v_match.group(2) else v_start
                if v_start == 1 and (v_end - v_start + 1) <= 5 and not has_ad_version:
                    return "Imagem Nova"
            return "Variação Imagem"

        if has_ad_version:
            return "Variação Vídeo"
        if v_match:
            v_start = int(v_match.group(1))
            v_end = int(v_match.group(2)) if v_match.group(2) else v_start
            if v_start == 1 and (v_end - v_start + 1) <= 5:
                return "Criativo Novo"
            return "Variação Vídeo"
        if re.search(r"\[AD\d+\s*[-aA]\s*(?:AD)?\s*\d+\]", task_name):
            return "Criativo Novo"
        return "Criativo Novo"

    # Fallback: se tem versão > V1, é variação
    if version:
        v_num = re.search(r"(\d+)", version)
        if v_num and int(v_num.group(1)) > 1:
            return "Variação Vídeo"
    return "Criativo Novo"


# === CORE ===
def collect_and_process(date_from, date_to):
    """Coleta dados e retorna dict de resultados."""
    print("[1/3] Carregando ClickUp...", flush=True)
    copy_raw = cached_cu_tasks(COPY_LIST, include_closed=True, ttl=3600)
    traf_raw = cached_cu_tasks(TRAFEGO_LIST, include_closed=False, ttl=3600)
    tasks = [parse_cu_task(t) for t in copy_raw + traf_raw if re.match(r"\s*\[", t["name"])]
    print(f"  {len(tasks)} tarefas\n", flush=True)

    print("[2/3] Carregando RedTrack...", flush=True)
    all_ags_raw, failed = fetch_rt_all_adgroups(date_from, date_to)

    print(f"\n[3/3] Processando...", flush=True)

    # Agregar por versão individual
    creative_data = defaultdict(lambda: {
        "cost": 0, "front_rev": 0, "vendas": 0, "nicho": None,
        "base_id": None, "version": None
    })

    stats = {"total_ag": 0, "sem_nicho": 0, "sem_base": 0}

    for ag in all_ags_raw:
        cost = float(ag.get("cost", 0))
        if cost < 1:
            continue
        stats["total_ag"] += 1
        cname = ag.get("campaign", "")
        nicho, fonte = parse_camp(cname)
        if not nicho:
            stats["sem_nicho"] += 1
            continue
        fr = float(ag.get("revenuetype2", 0)) + float(ag.get("revenuetype3", 0))
        vendas = int(ag.get("convtype1", 0))
        clean = clean_ag_name(ag.get("rt_adgroup", ""))
        base_id, version = extract_creative_id(clean)
        if not base_id:
            stats["sem_base"] += 1
            continue

        ver_key = version or "BASE"
        key = f"{nicho}_{base_id}_{ver_key}"
        creative_data[key]["cost"] += cost
        creative_data[key]["front_rev"] += fr
        creative_data[key]["vendas"] += vendas
        creative_data[key]["nicho"] = nicho
        creative_data[key]["base_id"] = base_id
        creative_data[key]["version"] = version

    # Classify + validate
    results = defaultdict(lambda: defaultdict(lambda: {
        "total": 0, "pre_validado": 0, "validado": 0,
        "total_spend": 0, "total_revenue": 0
    }))

    imaturos = 0
    sem_cw = 0

    for key, data in creative_data.items():
        if data["cost"] < SPEND_MINIMO:
            imaturos += 1
            continue
        base_id = data["base_id"]
        version = data["version"]
        nicho = data["nicho"]

        cw, ed, mt = find_copywriter(base_id, version, nicho, tasks)
        if cw == "REAPER":
            cw = "CASSIO"
        if not cw or cw in ("SEM ATRIBUIÇÃO", "N/A"):
            sem_cw += 1
            continue
        if cw == "DOUGLAS*":
            cw = "DOUGLAS (Gestor)"

        cat = classify_by_context(base_id, version, nicho, tasks)

        vendas = data["vendas"]
        cost = data["cost"]
        front_rev = data["front_rev"]
        roas_front = front_rev / cost if cost > 0 else 0

        results[cw][cat]["total"] += 1
        results[cw][cat]["total_spend"] += cost
        results[cw][cat]["total_revenue"] += front_rev

        if vendas >= 3 and roas_front >= ROAS_MIN:
            results[cw][cat]["pre_validado"] += 1
        if vendas >= 10 and roas_front >= ROAS_MIN:
            results[cw][cat]["validado"] += 1

    total_analisados = sum(d["total"] for cw in results for d in results[cw].values())
    print(f"  {len(creative_data)} criativos individuais | {total_analisados} analisados", flush=True)

    coverage = {
        "adgroups_total": stats["total_ag"],
        "criativos_total": len(creative_data),
        "analisados": total_analisados,
        "sem_nicho": stats["sem_nicho"],
        "sem_base": stats["sem_base"],
        "sem_cw": sem_cw,
        "imaturos": imaturos,
        "failed_campaigns": len(failed),
    }

    return results, coverage


# === PDF GENERATION ===
def generate_pdf(results, coverage, date_from, date_to):
    """Gera PDF com layout padrão IMPERA."""
    from reportlab.lib.pagesizes import A3, landscape
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    filename = os.path.join(OUTPUT_DIR, f"Assertividade_Copywriters_{date_from}_a_{date_to}.pdf")

    doc = SimpleDocTemplate(
        filename, pagesize=landscape(A3),
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
    )

    # Colors
    DARK_BLUE = colors.HexColor("#1B3A5C")
    MEDIUM_BLUE = colors.HexColor("#2C5F8A")
    LIGHT_BG = colors.HexColor("#F2F6FA")
    GREEN = colors.HexColor("#2E7D32")
    ORANGE = colors.HexColor("#E65100")
    RED = colors.HexColor("#C62828")
    GOLD = colors.HexColor("#B8860B")

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle("Title_IMPERA", fontSize=22, fontName="Helvetica-Bold",
                              textColor=DARK_BLUE, alignment=TA_CENTER, spaceAfter=6))
    styles.add(ParagraphStyle("Subtitle_IMPERA", fontSize=12, fontName="Helvetica",
                              textColor=MEDIUM_BLUE, alignment=TA_CENTER, spaceAfter=4))
    styles.add(ParagraphStyle("Section", fontSize=11, fontName="Helvetica-Bold",
                              textColor=DARK_BLUE, spaceBefore=12, spaceAfter=4))
    styles.add(ParagraphStyle("Body", fontSize=8, fontName="Helvetica", leading=10))
    styles.add(ParagraphStyle("Footer", fontSize=7, fontName="Helvetica-Oblique",
                              textColor=colors.gray, alignment=TA_CENTER))

    elements = []

    # === COVER ===
    elements.append(Spacer(1, 3 * cm))
    elements.append(Paragraph("RELATÓRIO DE ASSERTIVIDADE", styles["Title_IMPERA"]))
    elements.append(Paragraph("Por Copywriter e Categoria de Produção", styles["Subtitle_IMPERA"]))
    elements.append(Spacer(1, 0.5 * cm))
    elements.append(Paragraph(f"Período: {date_from} a {date_to}", styles["Subtitle_IMPERA"]))
    elements.append(Paragraph("IMPERA Produtos Naturais", styles["Subtitle_IMPERA"]))
    elements.append(Spacer(1, 1.5 * cm))

    # Metodologia
    elements.append(Paragraph("METODOLOGIA", styles["Section"]))
    metodo = [
        "• Cada versão individual que rodou em tráfego = 1 criativo testado",
        "• Atribuição ao copywriter que ESCREVEU aquela versão (match por range no ClickUp)",
        f"• Spend mínimo para considerar testado: R${SPEND_MINIMO}",
        "• ROAS Front = (revenuetype2 + revenuetype3) / cost",
        "",
        "<b>Critério Pré-validado+:</b> ≥3 vendas E ROAS Front ≥ 1.8 (passou no teste)",
        "<b>Critério Validado+:</b> ≥10 vendas E ROAS Front ≥ 1.8 (tração real, pronto para escala)",
        "",
        "Fonte: Super Cérebro de Tráfego V5 (Head de Tráfego)",
    ]
    for line in metodo:
        elements.append(Paragraph(line, styles["Body"]))

    # Cobertura
    elements.append(Spacer(1, 0.5 * cm))
    elements.append(Paragraph("COBERTURA DOS DADOS", styles["Section"]))
    cov_pct = (coverage["analisados"] / coverage["criativos_total"] * 100) if coverage["criativos_total"] else 0
    cov_data = [
        ["Adgroups com spend", str(coverage["adgroups_total"])],
        ["Criativos individuais (por versão)", str(coverage["criativos_total"])],
        ["Criativos analisados", f"{coverage['analisados']} ({cov_pct:.1f}%)"],
        ["Excluídos — Sem nicho (RT)", str(coverage["sem_nicho"])],
        ["Excluídos — Sem base_id", str(coverage["sem_base"])],
        ["Excluídos — Sem copywriter", str(coverage["sem_cw"])],
        ["Excluídos — Imaturos (spend < R$50)", str(coverage["imaturos"])],
        ["Campanhas falharam (429)", str(coverage["failed_campaigns"])],
    ]
    cov_table = Table(cov_data, colWidths=[8 * cm, 4 * cm])
    cov_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, LIGHT_BG]),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
    ]))
    elements.append(cov_table)
    elements.append(PageBreak())

    # === TABELA PRINCIPAL — Critério 1 ===
    elements.append(Paragraph("ASSERTIVIDADE — PRÉ-VALIDADO+ (≥3 vendas + ROAS ≥ 1.8)", styles["Section"]))

    header = ["Copywriter", "Global", "Testados", "Spend"] + CATEGORIAS
    rows = [header]

    sorted_cws = sorted(results.keys(), key=lambda x: (
        sum(d["pre_validado"] for d in results[x].values()) /
        max(sum(d["total"] for d in results[x].values()), 1)
    ), reverse=True)

    for cw in sorted_cws:
        t_all = sum(d["total"] for d in results[cw].values())
        p_all = sum(d["pre_validado"] for d in results[cw].values())
        s_all = sum(d["total_spend"] for d in results[cw].values())
        pct_g = f"{(p_all / t_all * 100):.1f}%" if t_all else "0%"
        row = [cw, pct_g, str(t_all), f"R${s_all:,.0f}"]
        for cat in CATEGORIAS:
            d = results[cw].get(cat)
            if not d or d["total"] == 0:
                row.append("—")
            else:
                pct = (d["pre_validado"] / d["total"] * 100)
                row.append(f"{pct:.0f}% ({d['pre_validado']}/{d['total']})")
        rows.append(row)

    # Totais empresa
    gt = sum(d["total"] for cw in results for d in results[cw].values())
    gp = sum(d["pre_validado"] for cw in results for d in results[cw].values())
    gs = sum(d["total_spend"] for cw in results for d in results[cw].values())
    total_row = ["EMPRESA", f"{(gp/gt*100):.1f}%" if gt else "0%", str(gt), f"R${gs:,.0f}"]
    for cat in CATEGORIAS:
        t = sum(results[cw].get(cat, {}).get("total", 0) for cw in results)
        p = sum(results[cw].get(cat, {}).get("pre_validado", 0) for cw in results)
        if t == 0:
            total_row.append("—")
        else:
            total_row.append(f"{(p/t*100):.0f}% ({p}/{t})")
    rows.append(total_row)

    col_widths = [3.5 * cm, 1.8 * cm, 1.8 * cm, 2.8 * cm] + [3.8 * cm] * len(CATEGORIAS)
    t1 = Table(rows, colWidths=col_widths)
    t1.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("BACKGROUND", (0, 0), (-1, 0), DARK_BLUE),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, LIGHT_BG]),
        ("BACKGROUND", (0, -1), (-1, -1), MEDIUM_BLUE),
        ("TEXTCOLOR", (0, -1), (-1, -1), colors.white),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(t1)
    elements.append(Spacer(1, 1 * cm))

    # === TABELA PRINCIPAL — Critério 2 ===
    elements.append(Paragraph("ASSERTIVIDADE — VALIDADO+ (≥10 vendas + ROAS ≥ 1.8)", styles["Section"]))

    rows2 = [header]
    sorted_cws2 = sorted(results.keys(), key=lambda x: (
        sum(d["validado"] for d in results[x].values()) /
        max(sum(d["total"] for d in results[x].values()), 1)
    ), reverse=True)

    for cw in sorted_cws2:
        t_all = sum(d["total"] for d in results[cw].values())
        v_all = sum(d["validado"] for d in results[cw].values())
        s_all = sum(d["total_spend"] for d in results[cw].values())
        pct_g = f"{(v_all / t_all * 100):.1f}%" if t_all else "0%"
        row = [cw, pct_g, str(t_all), f"R${s_all:,.0f}"]
        for cat in CATEGORIAS:
            d = results[cw].get(cat)
            if not d or d["total"] == 0:
                row.append("—")
            else:
                pct = (d["validado"] / d["total"] * 100)
                row.append(f"{pct:.0f}% ({d['validado']}/{d['total']})")
        rows2.append(row)

    # Totais empresa
    gv = sum(d["validado"] for cw in results for d in results[cw].values())
    total_row2 = ["EMPRESA", f"{(gv/gt*100):.1f}%" if gt else "0%", str(gt), f"R${gs:,.0f}"]
    for cat in CATEGORIAS:
        t = sum(results[cw].get(cat, {}).get("total", 0) for cw in results)
        v = sum(results[cw].get(cat, {}).get("validado", 0) for cw in results)
        if t == 0:
            total_row2.append("—")
        else:
            total_row2.append(f"{(v/t*100):.0f}% ({v}/{t})")
    rows2.append(total_row2)

    t2 = Table(rows2, colWidths=col_widths)
    t2.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("BACKGROUND", (0, 0), (-1, 0), DARK_BLUE),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, LIGHT_BG]),
        ("BACKGROUND", (0, -1), (-1, -1), MEDIUM_BLUE),
        ("TEXTCOLOR", (0, -1), (-1, -1), colors.white),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(t2)
    elements.append(PageBreak())

    # === DETALHAMENTO INDIVIDUAL ===
    elements.append(Paragraph("DETALHAMENTO POR COPYWRITER", styles["Section"]))

    for cw in sorted_cws:
        t_all = sum(d["total"] for d in results[cw].values())
        p_all = sum(d["pre_validado"] for d in results[cw].values())
        v_all = sum(d["validado"] for d in results[cw].values())
        s_all = sum(d["total_spend"] for d in results[cw].values())
        r_all = sum(d["total_revenue"] for d in results[cw].values())

        elements.append(Spacer(1, 0.4 * cm))
        elements.append(Paragraph(
            f"<b>{cw}</b> — {t_all} testados | Spend: R${s_all:,.0f} | Revenue: R${r_all:,.0f}",
            styles["Body"]
        ))

        detail_header = ["Categoria", "Testados", "Spend", "Pré-V+", "%Pré", "Val+", "%Val"]
        detail_rows = [detail_header]
        for cat in CATEGORIAS:
            d = results[cw].get(cat)
            if not d or d["total"] == 0:
                continue
            t, p, v = d["total"], d["pre_validado"], d["validado"]
            sp = d["total_spend"]
            detail_rows.append([
                cat, str(t), f"R${sp:,.0f}",
                str(p), f"{(p/t*100):.1f}%",
                str(v), f"{(v/t*100):.1f}%"
            ])
        # Total row
        detail_rows.append([
            "TOTAL", str(t_all), f"R${s_all:,.0f}",
            str(p_all), f"{(p_all/t_all*100):.1f}%" if t_all else "0%",
            str(v_all), f"{(v_all/t_all*100):.1f}%" if t_all else "0%"
        ])

        dw = [3.5 * cm, 1.5 * cm, 2.5 * cm, 1.3 * cm, 1.5 * cm, 1.3 * cm, 1.5 * cm]
        dt = Table(detail_rows, colWidths=dw)
        dt.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 7.5),
            ("BACKGROUND", (0, 0), (-1, 0), MEDIUM_BLUE),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, LIGHT_BG]),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E8EEF4")),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.2, colors.HexColor("#DDDDDD")),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
        ]))
        elements.append(dt)

    # === GLOSSÁRIO ===
    elements.append(Spacer(1, 1 * cm))
    elements.append(Paragraph("GLOSSÁRIO", styles["Section"]))
    glossario = [
        "<b>Criativo Novo:</b> AD com V1 (até V1-V5) — ideia original do copywriter",
        "<b>Variação Vídeo:</b> V2+, V6+, ou AD##V## — derivado de criativo existente",
        "<b>Imagem Nova:</b> IMG com V1 — criativo estático original",
        "<b>Variação Imagem:</b> IMG com V2+ — derivado de imagem existente",
        "<b>Ripagem:</b> Conteúdo ripado (CE/CY/CC/C##) ou marcado [RP]",
        "",
        "<b>ROAS Front:</b> (revenuetype2 + revenuetype3) / cost — receita front-end sobre custo",
        "<b>Spend mínimo:</b> R$50 — criativos abaixo disso são considerados imaturos (excluídos)",
        "<b>Pré-validado+:</b> ≥3 vendas + ROAS Front ≥ 1.8",
        "<b>Validado+:</b> ≥10 vendas + ROAS Front ≥ 1.8",
    ]
    for line in glossario:
        elements.append(Paragraph(line, styles["Body"]))

    # === ASSINATURA ===
    elements.append(Spacer(1, 1.5 * cm))
    elements.append(Paragraph("─" * 80, styles["Footer"]))
    elements.append(Paragraph(
        f"Desenvolvido por GPDR — Iago Almeida | Gestão de Performance e Dados em Resultados | IMPERA Produtos Naturais",
        styles["Footer"]
    ))
    elements.append(Paragraph(
        f"Assistido por Claude — Anthropic | v1.0 | Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        styles["Footer"]
    ))

    doc.build(elements)
    print(f"\n  📄 PDF: {filename}", flush=True)
    return filename


# === EXCEL GENERATION ===
def generate_excel(results, coverage, date_from, date_to):
    """Gera Excel com dados completos."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  ⚠️  openpyxl não instalado, pulando Excel")
        return None

    filename = os.path.join(OUTPUT_DIR, f"Assertividade_Copywriters_{date_from}_a_{date_to}.xlsx")
    wb = openpyxl.Workbook()

    # Colors
    header_fill = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
    data_font = Font(name="Calibri", size=9)
    total_fill = PatternFill(start_color="E8EEF4", end_color="E8EEF4", fill_type="solid")
    alt_fill = PatternFill(start_color="F2F6FA", end_color="F2F6FA", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # === Sheet 1: Resumo ===
    ws = wb.active
    ws.title = "Resumo"

    # Header
    headers = ["Copywriter", "Global Pré-V+", "Global Val+", "Testados", "Spend"]
    for cat in CATEGORIAS:
        headers.append(f"{cat} (Pré-V+)")
        headers.append(f"{cat} (Val+)")
        headers.append(f"{cat} (Total)")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    sorted_cws = sorted(results.keys(), key=lambda x: (
        sum(d["validado"] for d in results[x].values()) /
        max(sum(d["total"] for d in results[x].values()), 1)
    ), reverse=True)

    for row_idx, cw in enumerate(sorted_cws, 2):
        t_all = sum(d["total"] for d in results[cw].values())
        p_all = sum(d["pre_validado"] for d in results[cw].values())
        v_all = sum(d["validado"] for d in results[cw].values())
        s_all = sum(d["total_spend"] for d in results[cw].values())

        ws.cell(row=row_idx, column=1, value=cw)
        ws.cell(row=row_idx, column=2, value=p_all / t_all if t_all else 0)
        ws.cell(row=row_idx, column=3, value=v_all / t_all if t_all else 0)
        ws.cell(row=row_idx, column=4, value=t_all)
        ws.cell(row=row_idx, column=5, value=round(s_all, 2))

        col = 6
        for cat in CATEGORIAS:
            d = results[cw].get(cat, {"total": 0, "pre_validado": 0, "validado": 0})
            t = d["total"]
            ws.cell(row=row_idx, column=col, value=d["pre_validado"] / t if t else 0)
            ws.cell(row=row_idx, column=col + 1, value=d["validado"] / t if t else 0)
            ws.cell(row=row_idx, column=col + 2, value=t)
            col += 3

        # Formatting
        fill = alt_fill if row_idx % 2 == 0 else None
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.font = data_font
            cell.border = thin_border
            if fill:
                cell.fill = fill
            if c in (2, 3) or (c >= 6 and (c - 6) % 3 < 2):
                cell.number_format = '0.0%'
                cell.alignment = Alignment(horizontal="center")
            elif c in (4, 5) or (c >= 6 and (c - 6) % 3 == 2):
                cell.alignment = Alignment(horizontal="center")

    # Auto-width
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14

    # === Sheet 2: Dados brutos por copywriter ===
    ws2 = wb.create_sheet("Detalhamento")
    headers2 = ["Copywriter", "Categoria", "Testados", "Spend", "Revenue",
                "Pré-Validados", "%Pré", "Validados", "%Val"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    row_idx = 2
    for cw in sorted_cws:
        for cat in CATEGORIAS:
            d = results[cw].get(cat)
            if not d or d["total"] == 0:
                continue
            t, p, v = d["total"], d["pre_validado"], d["validado"]
            ws2.cell(row=row_idx, column=1, value=cw)
            ws2.cell(row=row_idx, column=2, value=cat)
            ws2.cell(row=row_idx, column=3, value=t)
            ws2.cell(row=row_idx, column=4, value=round(d["total_spend"], 2))
            ws2.cell(row=row_idx, column=5, value=round(d["total_revenue"], 2))
            ws2.cell(row=row_idx, column=6, value=p)
            ws2.cell(row=row_idx, column=7, value=p / t if t else 0)
            ws2.cell(row=row_idx, column=8, value=v)
            ws2.cell(row=row_idx, column=9, value=v / t if t else 0)

            for c in range(1, 10):
                cell = ws2.cell(row=row_idx, column=c)
                cell.font = data_font
                cell.border = thin_border
                if c in (7, 9):
                    cell.number_format = '0.0%'
                cell.alignment = Alignment(horizontal="center") if c > 1 else Alignment(horizontal="left")
            row_idx += 1

    for col in range(1, 10):
        ws2.column_dimensions[get_column_letter(col)].width = 15

    # === Sheet 3: Cobertura ===
    ws3 = wb.create_sheet("Cobertura")
    cov_rows = [
        ["Métrica", "Valor"],
        ["Período", f"{date_from} a {date_to}"],
        ["Adgroups com spend", coverage["adgroups_total"]],
        ["Criativos individuais", coverage["criativos_total"]],
        ["Analisados", coverage["analisados"]],
        ["Cobertura %", coverage["analisados"] / coverage["criativos_total"] if coverage["criativos_total"] else 0],
        ["Sem nicho (RT)", coverage["sem_nicho"]],
        ["Sem base_id", coverage["sem_base"]],
        ["Sem copywriter", coverage["sem_cw"]],
        ["Imaturos (< R$50)", coverage["imaturos"]],
        ["Campanhas falharam", coverage["failed_campaigns"]],
        ["Spend mínimo", f"R${SPEND_MINIMO}"],
        ["ROAS mínimo", ROAS_MIN],
    ]
    for r, row in enumerate(cov_rows, 1):
        for c, val in enumerate(row, 1):
            cell = ws3.cell(row=r, column=c, value=val)
            cell.font = header_font if r == 1 else data_font
            if r == 1:
                cell.fill = header_fill
            cell.border = thin_border
            if r == 6 and c == 2:
                cell.number_format = '0.0%'
    ws3.column_dimensions["A"].width = 25
    ws3.column_dimensions["B"].width = 20

    wb.save(filename)
    print(f"  📊 Excel: {filename}", flush=True)
    return filename


# === MAIN ===
def main():
    date_from, date_to = parse_dates()
    print(f"\n{'='*70}")
    print(f"  RELATÓRIO DE ASSERTIVIDADE — IMPERA")
    print(f"  Período: {date_from} a {date_to}")
    print(f"{'='*70}\n")

    results, coverage = collect_and_process(date_from, date_to)

    # Console output
    total_analisados = coverage["analisados"]
    print(f"\n{'═'*90}")
    print(f"  ASSERTIVIDADE — Abril/2026 | {total_analisados} criativos analisados")
    print(f"{'═'*90}")

    CATS = CATEGORIAS
    print(f"\n  CRITÉRIO: Validado+ (≥10 vendas + ROAS Front ≥ 1.8)")
    print(f"  {'─'*85}")
    print(f"  {'Copywriter':<18} {'Global':<8} {'N':<5} │ ", end="")
    print("  ".join(f"{c[:10]:<12}" for c in CATS))
    print(f"  {'─'*85}")

    sorted_cws = sorted(results.keys(), key=lambda x: (
        sum(d["validado"] for d in results[x].values()) /
        max(sum(d["total"] for d in results[x].values()), 1)
    ), reverse=True)

    for cw in sorted_cws:
        t_all = sum(d["total"] for d in results[cw].values())
        v_all = sum(d["validado"] for d in results[cw].values())
        pct = (v_all / t_all * 100) if t_all else 0
        line = f"  {cw:<18} {pct:<8.1f} {t_all:<5} │ "
        for cat in CATS:
            d = results[cw].get(cat)
            if not d or d["total"] == 0:
                line += f"{'—':<14}"
            else:
                p = (d["validado"] / d["total"] * 100)
                cell = f"{p:.0f}% ({d['validado']}/{d['total']})"
                line += f"{cell:<14}"
        print(line)

    # Generate outputs
    print(f"\nGerando documentos...", flush=True)
    generate_pdf(results, coverage, date_from, date_to)
    generate_excel(results, coverage, date_from, date_to)

    print(f"\n{'─'*70}")
    print(f"  GPDR — Iago Almeida | IMPERA Produtos Naturais")
    print(f"  {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    print(f"{'─'*70}\n")


if __name__ == "__main__":
    main()
